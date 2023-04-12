from openpyxl import load_workbook


class Extractor:
    def __init__(self, xlsx_file_name: str):
        self.work_dict = {}
        xlsx_load = load_workbook(xlsx_file_name + '.xlsx')
        sheets_name_list = list(xlsx_load.sheetnames)
        for sheet_name in sheets_name_list:
            work_sheet = list(xlsx_load[sheet_name])
            self.work_dict[sheet_name] = {}
            next_question = False
            for row in work_sheet:
                for cell in row:
                    if isinstance(cell.value, int):
                        next_question = True
                    elif next_question:
                        next_question = False
                        question = cell.value
                        self.work_dict[sheet_name][cell.value] = []
                    elif cell.value is not None:
                        self.work_dict[sheet_name][question].append(cell.value)
