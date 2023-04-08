from openpyxl import load_workbook

xlsx_file = load_workbook('exemple.xlsx')

for sheetname in xlsx_file.sheetnames:
    worksheet = xlsx_file[sheetname]
    for row in list(worksheet):
        for cell in row:
            print(worksheet, cell.value)
